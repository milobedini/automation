import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# CREATING PIVOT TABLE

# xlsx is excel
df = pd.read_excel("supermarket_sales.xlsx")

# Want Gender, Product Line and Total columns.
df = df[["Gender", "Product line", "Total"]]

# A pivot table is a table of grouped values that aggregates the individual items of a more extensive table within one or more discrete categories.

# Goal is to see how much each gender spends in each product line.
# So gender the index, and product line the columns.
pivot_table = df.pivot_table(index="Gender", columns="Product line", values="Total", aggfunc="sum").round(0)

# Create file, with sheet called report and with the data starting in row 4.
pivot_table.to_excel("pivot_table.xlsx", "Report", startrow=4)

# CREATING BAR CHART
# pip install openpyxl

wb = load_workbook("pivot_table.xlsx")
sheet = wb["Report"]

min_column = wb.active.min_column  # prints 1 (A)
max_column = wb.active.max_column  # prints 7 (G)
min_row = wb.active.min_row  # prints 5
max_row = wb.active.max_row  # prints 7

barchart = BarChart()

data = Reference(
    sheet,
    # Account for data starting in B
    min_col=min_column + 1,
    max_col=max_column,
    min_row=min_row,
    max_row=max_row,
)

# NB data does include the header/labels but the categories do not.

categories = Reference(
    sheet,
    min_col=min_column,
    # Only want Column A
    max_col=min_column,
    # Starts from row 6
    min_row=min_row + 1,
    max_row=max_row,
)

# Create in B12
barchart.add_data(data, titles_from_data=True)
barchart.set_categories(categories)
barchart.title = "Sales by gender and product line"
barchart.style = 5

sheet.add_chart(barchart, "B12")

wb.save("barchart.xlsx")

# USE EXCEL FORMULAE IN PYTHON

# =SUM(B6:B7), electronics
# For one cell
# sheet["B8"] = "=SUM(B6:B7)"
# sheet["B8"].style = "Currency"

# We often want to mass create formulae however.
# Want to sum columns B-G downwards, using rows 6 and 7.

for i in range(min_column + 1, max_column + 1):
    column = get_column_letter(i)
    sheet[f"{column}{max_row+1}"] = f"=SUM({column}{min_row+1}:{column}{max_row})"
    sheet[f"{column}{max_row+1}"].style = "Currency"

# FORMATTING

sheet["A1"] = "Sales Report"
sheet["A2"] = "January"
sheet["A1"].font = Font("Arial", bold=True, size=20)
sheet["A2"].font = Font("Arial", italic=True, size=16)


wb.save("formulae.xlsx")
